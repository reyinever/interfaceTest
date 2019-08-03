from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, fonts
import time


class ParseExcel(object):
    """解析excel文件"""

    def __init__(self):
        self.workbook = None
        self.excelfilepath = None

    def loadWorkBook(self, excel_file_path):
        # 加载excel文件到内存
        try:
            self.workbook = load_workbook(excel_file_path)
            self.excelfilepath = excel_file_path
        except Exception as e:
            raise e
        return self.workbook

    def getSheetByName(self, sheet_name):
        # 根据表名获取sheet对象
        try:
            sheet = self.workbook.get_sheet_by_name(sheet_name)
            # print(sheet.title)
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheet_index):
        # 根据sheet索引获取sheet对象
        try:
            sheet_name = self.workbook.get_sheet_names()[sheet_index]
            sheet = self.workbook.get_sheet_by_name(sheet_name)
            return sheet
        except Exception as e:
            raise e

    def getRowsNumber(self, sheet):
        # 获取sheet的最大行号
        return sheet.max_row

    def getColsNumber(self, sheet):
        # 获取sheet的最大列号
        return sheet.max_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中rowNo行的对象
        try:
            rows = []
            for row in sheet.iter_rows():
                rows.append(row)
            return rows[rowNo - 1]
        except Exception as e:
            raise e

    def getCol(self, sheet, colNo):
        # 获取sheet中rowCol列的对象
        try:
            cols = []
            for col in sheet.iter_cols():
                cols.append(col)
            return cols[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 根据单元格的位置获取单元格的值
        if coordinate is not None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo).value
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient coordinates of cell !')

    def getCellOfObject(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 获取某个单元格的对象
        if coordinate is not None:
            try:
                return sheet[coordinate]
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient coordinates of cell !')

    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colNo=None, font_color=None):
        # 向指定单元格中写入内容
        # 把字体颜色字符串变成大写
        if font_color is not None:
            try:
                font_color = font_color.strip().upper()
                # print(fontcolor)
            except Exception as e:
                raise e
        if coordinate is not None:
            try:
                sheet[coordinate].value = content
                if font_color:
                    exec('sheet[coordinate].font=Font(color=colors.' + font_color + ')')
                self.workbook.save(self.excelfilepath)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = content
                if font_color:
                    exec('sheet.cell(row=rowNo,column=colNo).font=Font(color=colors.' + font_color + ')')
                self.workbook.save(self.excelfilepath)
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient coordinates of cell !')

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 向指定单元格中写入当前时间
        if coordinate is not None:
            try:
                current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                sheet[coordinate].value = current_time
                self.workbook.save(self.excelfilepath)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                sheet.cell(row=rowNo, column=colNo).value = current_time
                self.workbook.save(self.excelfilepath)
            except Exception as e:
                raise e
        else:
            raise Exception('Insufficient coordinate of cell !')
